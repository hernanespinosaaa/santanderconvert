#!/usr/bin/env python3
"""
Conversor Santander PDF → Excel
Streamlit app · Parser basado en validación matemática de saldos
"""

import re
import io
from datetime import datetime

import pdfplumber
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Constantes de color Excel ─────────────────────────────────────────────────
C_HEADER_BG  = "CC0000"
C_HEADER_FG  = "FFFFFF"
C_SUBHEADER  = "F2F2F2"
C_DEBITO     = "FFF0F0"
C_CREDITO    = "F0FFF0"
C_IMPUESTO   = "FFF8E1"
C_TOTAL_BG   = "EEEEEE"
C_SALDO_INIT = "EEF2FF"

THIN = Side(style="thin", color="CCCCCC")
BRD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── Regex globales ────────────────────────────────────────────────────────────
RE_MONTO_AR  = re.compile(r"-?\$\s*[\d.]+,\d{2}")
RE_MONTO_USD = re.compile(r"U\$S\s*[\d.]+,\d{2}")
RE_FECHA     = re.compile(r"^\d{2}/\d{2}/\d{2}$")
RE_CBU       = re.compile(r"N[°ºoO]?\s*([\d\-/]+)\s+CBU:\s*(\d+)", re.IGNORECASE)
RE_COMP      = re.compile(r"^\d{5,9}$")
RE_PAG       = re.compile(r"^\d+\s*-\s*\d+$")

STOP_PALABRAS = (
    "detalle impositivo", "legales", "intercambio de información",
    "movimientos en dólares", "saldo total",
)

# ── Helpers ───────────────────────────────────────────────────────────────────
def parse_monto(s: str) -> float | None:
    if not s:
        return None
    s = s.strip()
    neg = s.startswith("-")
    s = re.sub(r"[^\d,.]", "", s)
    if not s:
        return None
    if re.search(r",\d{2}$", s):
        s = s.replace(".", "").replace(",", ".")
    try:
        v = float(s)
        return -v if neg else v
    except ValueError:
        return None


def categorizar(d: str) -> str:
    d = d.lower()
    if "saldo inicial" in d:                           return "Saldo Inicial"
    if "haberes" in d or "sueldo" in d:                return "Sueldos y haberes"
    if "honorario" in d or "/ hon" in d:               return "Honorarios"
    if "alquiler" in d or "/ alq" in d:                return "Alquileres"
    if "afip" in d or "imp.afp" in d:                  return "Impuestos AFIP"
    if "ley 25.413" in d:                              return "Imp. débitos/créditos"
    if "ii bb" in d or "iibb" in d:                    return "IIBB"
    if "iva" in d:                                     return "IVA"
    if "fondos comunes" in d or "rescate" in d:        return "FCI"
    if "suscripcion" in d or "suscripción" in d:       return "FCI"
    if "debin" in d:                                   return "DEBIN"
    if "seguro" in d or "/ seg" in d:                  return "Seguros"
    if "comision" in d:                                return "Comisiones bancarias"
    if "prev.social" in d or "/ cuo" in d:             return "Previsión social"
    if "transferencia" in d or "transf" in d:          return "Transferencias"
    if "acreditacion" in d or "acreditación" in d:     return "Acreditaciones"
    if "cheque" in d or "echeq" in d:                  return "Cheques"
    if "liquidacion titulos" in d:                     return "Títulos/Bonos"
    if "pago de servicios" in d or "snp" in d:         return "Servicios"
    if "pago tarjeta" in d:                            return "Tarjeta de crédito"
    return "Otros"


# ── Extracción de info global ─────────────────────────────────────────────────
def extraer_info_global(lineas: list[str]) -> dict:
    info = {"cuit": "", "desde": "", "hasta": "", "razon_social": ""}
    for l in lineas:
        ls = l.strip()
        if not info["cuit"]:
            m = re.search(r"CUIT[:\s]+([\d\-]+)", ls)
            if m:
                info["cuit"] = m.group(1)
        if not info["desde"]:
            m = re.search(r"Desde:\s*(\d{2}/\d{2}/\d{2})", ls)
            if m:
                info["desde"] = m.group(1)
        if not info["hasta"]:
            m = re.search(r"Hasta:\s*(\d{2}/\d{2}/\d{2})", ls)
            if m:
                info["hasta"] = m.group(1)
        if not info["razon_social"] and info["cuit"]:
            if (re.match(r"^[A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s\.\-&,]+$", ls)
                    and 4 < len(ls) < 60
                    and ls not in ("EXTRACTO DE CUENTA", "CUENTA CORRIENTE",
                                   "BANCO SANTANDER", "RESUMEN DE CUENTA")):
                info["razon_social"] = ls
    return info


# ── Extracción multicuenta ────────────────────────────────────────────────────
def extraer_cuentas_del_pdf(pdf_file) -> tuple[dict, list[dict]]:
    """
    Devuelve (info_global, lista_de_cuentas).
    Cada cuenta: {"nro", "cbu", "moneda", "lineas"}
    """
    data = io.BytesIO(pdf_file.read())
    with pdfplumber.open(data) as pdf:
        lineas_raw = []
        for page in pdf.pages:
            lineas_raw.extend((page.extract_text() or "").splitlines())

    info_global = extraer_info_global(lineas_raw)

    cuentas = {}
    orden = []
    cuenta_act = None
    capturando = False

    for l in lineas_raw:
        ls = l.strip()
        ll = ls.lower()

        # Parar en secciones no contables
        if any(sw in ll for sw in STOP_PALABRAS):
            capturando = False
            cuenta_act = None
            continue

        # Nuevo encabezado de cuenta
        m = RE_CBU.search(ls)
        if m and "Fecha" not in ls:
            nro = m.group(1)
            cbu = m.group(2)
            # Detectar moneda por el contexto cercano (U$S o no)
            moneda = "U$S" if "U$S" in ls or "dólares" in ll else "$"
            cuenta_act = nro
            capturando = True
            if nro not in cuentas:
                cuentas[nro] = {"nro": nro, "cbu": cbu, "moneda": moneda, "lineas": []}
                orden.append(nro)
            continue

        if not capturando or not cuenta_act:
            continue

        # Filtros de ruido
        if (RE_PAG.match(ls)
                or ("Fecha" in ls and "Comprobante" in ls)
                or ls.lower().startswith("* salvo")
                or re.match(r"^total\s*(u\$s|\$|$)", ll)
                or not ls):
            if re.match(r"^total\s*(u\$s|\$|$)", ll):
                capturando = False
                cuenta_act = None
            continue

        cuentas[cuenta_act]["lineas"].append(ls)

    return info_global, [cuentas[n] for n in orden]


# ── Parser de movimientos ─────────────────────────────────────────────────────
def procesar_lineas(lineas: list[str], moneda: str = "$") -> list[dict]:
    """
    Convierte líneas de texto a lista de movimientos.

    Patrón real del PDF Santander:
      LÍNEA A: [DD/MM/AA] [comprobante] descripcion  $importe  $saldo
      LÍNEA B: beneficiario/referencia  (sin montos, opcional)
      LÍNEA C: DD/MM/AA sola            ← fecha del movimiento de LÍNEA A cuando no viene en ella

    La fecha suelta que aparece DESPUÉS de un bloque de movimiento
    pertenece a ese movimiento, no al siguiente.

    Clasificación débito/crédito por validación matemática:
      saldo_anterior + importe = saldo_posterior
      Si la diferencia es positiva → crédito; negativa → débito.
    """
    RE_M = RE_MONTO_AR if moneda == "$" else RE_MONTO_USD
    movimientos: list[dict] = []
    saldo_actual: float | None = None
    i = 0

    # ── Saldo inicial ────────────────────────────────────────────────────────
    for j, l in enumerate(lineas):
        if "saldo inicial" in l.lower():
            ms = RE_M.findall(l)
            if ms:
                saldo_actual = parse_monto(ms[-1])
                mf = RE_FECHA.search(l)
                movimientos.append({
                    "fecha":        mf.group(0) if mf else "",
                    "comprobante":  "",
                    "descripcion":  "Saldo Inicial",
                    "debito":       None,
                    "credito":      None,
                    "saldo":        saldo_actual,
                    "categoria":    "Saldo Inicial",
                    "sin_fecha":    not bool(mf),
                })
                i = j + 1
            break

    # ── Movimientos ──────────────────────────────────────────────────────────
    while i < len(lineas):
        l = lineas[i].strip()
        i += 1

        if not l or RE_PAG.match(l):
            continue
        if "no tenés movimientos" in l.lower():
            continue
        if re.match(r"^total\b", l.lower()):
            break
        # Fecha suelta que no pertenece a ningún bloque en curso → ignorar
        if RE_FECHA.match(l):
            continue

        montos_raw = RE_M.findall(l)
        if not montos_raw:
            # Texto de beneficiario/referencia: agregar al movimiento anterior
            if movimientos and movimientos[-1]["categoria"] != "Saldo Inicial":
                movimientos[-1]["descripcion"] += " | " + l
                movimientos[-1]["categoria"] = categorizar(movimientos[-1]["descripcion"])
            continue

        # ── Línea con montos: es un movimiento ──────────────────────────────
        sin_m  = RE_M.sub("", l).strip()
        tokens = sin_m.split()
        fecha = comp = ""
        desc_t: list[str] = []
        for t in tokens:
            if not fecha and RE_FECHA.match(t):
                fecha = t
            elif not comp and RE_COMP.match(t):
                comp = t
            else:
                desc_t.append(t)
        desc = " ".join(desc_t).strip()

        # Mirar las líneas siguientes para capturar beneficiario y/o fecha suelta
        # Patrón: [beneficiario_texto?] [fecha_sola?]
        # La fecha suelta cierra el bloque del movimiento actual.
        j = i
        while j < len(lineas):
            sig = lineas[j].strip()
            if not sig or RE_PAG.match(sig):
                j += 1
                continue
            if re.match(r"^total\b", sig.lower()):
                break
            if RE_M.findall(sig):
                # Siguiente movimiento con montos → salir sin consumir
                break
            if RE_FECHA.match(sig):
                # Fecha suelta: asignar al movimiento actual si no tenía
                if not fecha:
                    fecha = sig
                j += 1  # consumir la línea de fecha
                break
            # Es texto de beneficiario/referencia
            desc = (desc + " | " + sig) if desc else sig
            j += 1
            # Después del beneficiario puede venir una fecha suelta: continuar el loop
        i = j

        # ── Parsear importes ─────────────────────────────────────────────────
        vals = [parse_monto(m) for m in montos_raw if parse_monto(m) is not None]
        debito = credito = saldo = None

        if len(vals) >= 2:
            importe = vals[-2]
            saldo   = vals[-1]
            abs_imp = abs(importe)

            # 1. Validación matemática (principal)
            if saldo_actual is not None:
                dif = round(saldo - saldo_actual, 2)
                if abs(abs(dif) - round(abs_imp, 2)) < 0.02:
                    credito = abs_imp if dif > 0 else None
                    debito  = abs_imp if dif <= 0 else None
                else:
                    # 2. Fallback por descripción / signo
                    desc_l = desc.lower()
                    if "rescate" in desc_l:
                        credito = abs_imp
                    elif "suscripcion" in desc_l or "suscripción" in desc_l:
                        debito = abs_imp
                    elif importe < 0:
                        debito = abs_imp
                    elif any(k in desc_l for k in (
                        "recibida", "credi transf", "crédito transf",
                        "credito transf", "liquidacion titulos publicos credi",
                    )):
                        credito = abs_imp
                    else:
                        debito = abs_imp
            else:
                debito = abs_imp  # sin saldo anterior, asumir débito

        elif vals:
            saldo = vals[0]

        if saldo is not None:
            saldo_actual = saldo

        movimientos.append({
            "fecha":       fecha,
            "comprobante": comp,
            "descripcion": desc,
            "debito":      debito,
            "credito":     credito,
            "saldo":       saldo,
            "categoria":   categorizar(desc),
            "sin_fecha":   fecha == "",
        })

    return movimientos


# ── Creación del Excel en memoria ─────────────────────────────────────────────
def hdr(cell, bg=C_HEADER_BG, fg=C_HEADER_FG, bold=True, sz=10):
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=sz)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BRD


def crear_excel(info: dict, movimientos: list[dict], moneda: str = "$") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title        = "Movimientos"
    ws.freeze_panes = "A5"

    simbolo = "$" if moneda == "$" else "U$S"

    # Encabezado
    ws.merge_cells("A1:H1")
    ws["A1"].value = (
        f"Resumen  {info.get('razon_social', '')}  ·  "
        f"{info.get('desde', '')} al {info.get('hasta', '')}"
    )
    hdr(ws["A1"], sz=12)
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = (
        f"Cta Nº {info.get('nro_cuenta', '')}  ·  "
        f"CBU: {info.get('cbu', '')}  ·  CUIT: {info.get('cuit', '')}"
    )
    s.font      = Font(name="Arial", size=9, color="555555")
    s.fill      = PatternFill("solid", fgColor=C_SUBHEADER)
    s.alignment = Alignment(horizontal="center")
    s.border    = BRD
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 4

    # Encabezado de columnas
    cols_hdr = ["Fecha", "Comprobante", "Descripción", "Categoría",
                f"Débito ({simbolo})", f"Crédito ({simbolo})", f"Saldo ({simbolo})", ""]
    for c, h in enumerate(cols_hdr, 1):
        cell = ws.cell(row=4, column=c, value=h)
        hdr(cell, bg="8B0000" if c in (5, 6, 7) else C_HEADER_BG)
    ws.row_dimensions[4].height = 16

    for idx, w in enumerate([10, 12, 50, 22, 15, 15, 15, 2], 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    num_fmt = '#,##0.00;[Red](#,##0.00);-'

    # Filas de movimientos
    fila = 5
    for mov in movimientos:
        bg = "FFFFFF"
        cat = mov["categoria"]
        if cat == "Saldo Inicial":
            bg = C_SALDO_INIT
        elif mov.get("sin_fecha"):
            bg = "FFF3CD"
        elif mov["debito"] and not mov["credito"]:
            bg = C_DEBITO
        elif mov["credito"] and not mov["debito"]:
            bg = C_CREDITO
        if cat in ("Imp. débitos/créditos", "IVA", "IIBB"):
            bg = C_IMPUESTO

        valores = [
            mov["fecha"], mov["comprobante"], mov["descripcion"], cat,
            mov["debito"], mov["credito"], mov["saldo"],
        ]
        for c, v in enumerate(valores, 1):
            cell        = ws.cell(row=fila, column=c, value=v)
            cell.font   = Font(name="Arial", size=9)
            cell.fill   = PatternFill("solid", fgColor=bg)
            cell.border = BRD
            if c in (5, 6, 7) and v is not None:
                cell.number_format = num_fmt
                cell.alignment     = Alignment(horizontal="right")
        fila += 1

    # Fila de totales
    for c in range(1, 9):
        ws.cell(row=fila, column=c).fill   = PatternFill("solid", fgColor=C_TOTAL_BG)
        ws.cell(row=fila, column=c).border = BRD
    ws.cell(row=fila, column=4, value="TOTALES").font = Font(name="Arial", bold=True, size=9)
    for c, col in [(5, "E"), (6, "F")]:
        cell               = ws.cell(row=fila, column=c)
        cell.value         = f"=SUM({col}5:{col}{fila - 1})"
        cell.font          = Font(name="Arial", bold=True, size=9)
        cell.number_format = num_fmt
        cell.alignment     = Alignment(horizontal="right")

    # ── Hoja Por Categoría ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Por Categoría")
    for idx, w in enumerate([28, 16, 16, 8], 1):
        ws2.column_dimensions[get_column_letter(idx)].width = w
    ws2.merge_cells("A1:D1")
    ws2["A1"].value = "Resumen por Categoría"
    hdr(ws2["A1"], sz=11)
    for c, h in enumerate(["Categoría", f"Débitos ({simbolo})", f"Créditos ({simbolo})", "Movs."], 1):
        hdr(ws2.cell(row=2, column=c, value=h))

    cats: dict = {}
    for m in movimientos:
        if m["categoria"] == "Saldo Inicial":
            continue
        cats.setdefault(m["categoria"], {"d": 0.0, "c": 0.0, "n": 0})
        cats[m["categoria"]]["d"] += m["debito"]  or 0
        cats[m["categoria"]]["c"] += m["credito"] or 0
        cats[m["categoria"]]["n"] += 1

    r = 3
    for cat, v in sorted(cats.items()):
        ws2.cell(row=r, column=1, value=cat).font   = Font(name="Arial", size=9)
        ws2.cell(row=r, column=1).border            = BRD
        for c, k in [(2, "d"), (3, "c"), (4, "n")]:
            cell               = ws2.cell(row=r, column=c, value=v[k])
            cell.font          = Font(name="Arial", size=9)
            cell.alignment     = Alignment(horizontal="right")
            cell.border        = BRD
            if c in (2, 3):
                cell.number_format = num_fmt
        r += 1

    # ── Hoja Info ─────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Info")
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 40
    ws3.merge_cells("A1:B1")
    ws3["A1"].value = "Datos del Extracto"
    hdr(ws3["A1"])

    saldo_final = movimientos[-1]["saldo"] if movimientos else ""
    campos = [
        ("Razón Social",  info.get("razon_social", "")),
        ("CUIT",          info.get("cuit", "")),
        ("N° Cuenta",     info.get("nro_cuenta", "")),
        ("CBU",           info.get("cbu", "")),
        ("Moneda",        moneda),
        ("Período Desde", info.get("desde", "")),
        ("Período Hasta", info.get("hasta", "")),
        ("Saldo Inicial", movimientos[0]["saldo"] if movimientos else ""),
        ("Saldo Final",   saldo_final),
        ("Total Movim.",  len(movimientos) - 1),   # excluir saldo inicial
        ("Generado",      datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]
    for idx, (k, v) in enumerate(campos, 2):
        ws3.cell(row=idx, column=1, value=k).font = Font(name="Arial", bold=True, size=9)
        cell      = ws3.cell(row=idx, column=2, value=v)
        cell.font = Font(name="Arial", size=9)
        if k in ("Saldo Inicial", "Saldo Final") and isinstance(v, float):
            cell.number_format = num_fmt

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Procesamiento completo de un PDF ──────────────────────────────────────────
def procesar_pdf(pdf_file) -> list[dict]:
    """
    Retorna lista de resultados, uno por cuenta con movimientos.
    Cada resultado: {"info": {...}, "movimientos": [...], "moneda": "..."}
    """
    info_global, cuentas = extraer_cuentas_del_pdf(pdf_file)
    resultados = []

    for cta in cuentas:
        movimientos = procesar_lineas(cta["lineas"], cta["moneda"])
        movs_reales = [m for m in movimientos if m["categoria"] != "Saldo Inicial"]
        if not movs_reales:
            continue

        info_cta = {
            **info_global,
            "nro_cuenta": cta["nro"],
            "cbu":        cta["cbu"],
        }
        resultados.append({
            "info":        info_cta,
            "movimientos": movimientos,
            "moneda":      cta["moneda"],
        })

    return resultados


# ══════════════════════════════════════════════════════════════════════════════
# UI Streamlit
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Santander → Excel",
    page_icon="🏦",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ── CSS personalizado ─────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;500;600&display=swap');

html, body, [class*="css"] {
    font-family: 'IBM Plex Sans', sans-serif;
}

/* Fondo general */
.stApp {
    background: #f8f6f2;
}

/* Ocultar elementos default de Streamlit */
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 2rem; padding-bottom: 3rem; max-width: 780px; }

/* Header personalizado */
.app-header {
    background: #CC0000;
    border-radius: 12px;
    padding: 28px 32px 22px;
    margin-bottom: 28px;
    position: relative;
    overflow: hidden;
}
.app-header::before {
    content: '';
    position: absolute;
    top: -40px; right: -40px;
    width: 160px; height: 160px;
    background: rgba(255,255,255,0.06);
    border-radius: 50%;
}
.app-header::after {
    content: '';
    position: absolute;
    bottom: -20px; right: 60px;
    width: 80px; height: 80px;
    background: rgba(255,255,255,0.04);
    border-radius: 50%;
}
.app-header h1 {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1.6rem;
    font-weight: 600;
    color: #fff;
    margin: 0 0 4px;
    letter-spacing: -0.5px;
}
.app-header p {
    font-size: 0.85rem;
    color: rgba(255,255,255,0.72);
    margin: 0;
    font-weight: 300;
}

/* Zona de upload */
.upload-zone {
    background: #fff;
    border: 2px dashed #ddd;
    border-radius: 12px;
    padding: 8px 16px 16px;
    margin-bottom: 16px;
    transition: border-color 0.2s;
}
.upload-zone:hover { border-color: #CC0000; }

/* Tarjeta de resultado */
.result-card {
    background: #fff;
    border-radius: 12px;
    border: 1px solid #e8e4df;
    padding: 20px 24px;
    margin-bottom: 14px;
    position: relative;
}
.result-card-header {
    display: flex;
    align-items: flex-start;
    justify-content: space-between;
    margin-bottom: 12px;
}
.result-badge {
    background: #CC0000;
    color: #fff;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.72rem;
    font-weight: 600;
    padding: 3px 10px;
    border-radius: 20px;
    letter-spacing: 0.5px;
}
.result-badge-usd {
    background: #1a5276;
}
.result-account {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1.05rem;
    font-weight: 600;
    color: #1a1a1a;
    margin-bottom: 2px;
}
.result-meta {
    font-size: 0.78rem;
    color: #888;
}
.stat-row {
    display: flex;
    gap: 16px;
    margin-top: 10px;
    padding-top: 10px;
    border-top: 1px solid #f0ece8;
}
.stat-item { flex: 1; }
.stat-label {
    font-size: 0.7rem;
    color: #aaa;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 2px;
}
.stat-value {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.88rem;
    font-weight: 600;
    color: #1a1a1a;
}
.stat-value.red { color: #CC0000; }
.stat-value.green { color: #1e7e34; }

/* Botón de descarga */
.stDownloadButton button {
    background: #CC0000 !important;
    color: #fff !important;
    border: none !important;
    border-radius: 8px !important;
    font-family: 'IBM Plex Sans', sans-serif !important;
    font-weight: 500 !important;
    font-size: 0.85rem !important;
    padding: 8px 20px !important;
    width: 100% !important;
    margin-top: 10px !important;
    transition: background 0.15s !important;
}
.stDownloadButton button:hover {
    background: #a80000 !important;
}

/* Botón procesar */
.stButton button {
    background: #1a1a1a !important;
    color: #fff !important;
    border: none !important;
    border-radius: 8px !important;
    font-family: 'IBM Plex Sans', sans-serif !important;
    font-weight: 500 !important;
    font-size: 0.9rem !important;
    padding: 10px 28px !important;
    width: 100% !important;
    transition: background 0.15s !important;
}
.stButton button:hover { background: #333 !important; }

/* Alertas */
.stAlert { border-radius: 8px !important; }

/* Spinner */
.stSpinner > div { border-top-color: #CC0000 !important; }

/* File uploader */
[data-testid="stFileUploader"] {
    background: transparent !important;
}

/* Expander */
.streamlit-expanderHeader {
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 0.82rem !important;
    font-weight: 600 !important;
    color: #555 !important;
    background: #f8f6f2 !important;
    border-radius: 8px !important;
}

/* Legend */
.legend-row {
    display: flex;
    gap: 14px;
    flex-wrap: wrap;
    margin-top: 6px;
    font-size: 0.75rem;
    color: #666;
}
.legend-dot {
    display: inline-block;
    width: 10px; height: 10px;
    border-radius: 2px;
    margin-right: 4px;
    vertical-align: middle;
}
</style>
""", unsafe_allow_html=True)

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="app-header">
    <h1>Santander → Excel</h1>
    <p>Convertí extractos PDF a planillas Excel · Multicuenta · Validación matemática de saldos</p>
</div>
""", unsafe_allow_html=True)

# ── Estado de sesión ──────────────────────────────────────────────────────────
if "resultados_por_archivo" not in st.session_state:
    st.session_state.resultados_por_archivo = {}

# ── Upload ────────────────────────────────────────────────────────────────────
st.markdown('<div class="upload-zone">', unsafe_allow_html=True)
uploaded_files = st.file_uploader(
    "Arrastrá uno o más extractos PDF",
    type=["pdf"],
    accept_multiple_files=True,
    label_visibility="visible",
)
st.markdown("</div>", unsafe_allow_html=True)

if not uploaded_files:
    st.session_state.resultados_por_archivo = {}
    st.markdown("""
    <div style="text-align:center;padding:32px 0 16px;color:#aaa;font-size:0.82rem;">
        Subí el PDF del resumen mensual de Santander.<br>
        Soporta múltiples cuentas dentro del mismo archivo.
    </div>
    """, unsafe_allow_html=True)

if uploaded_files:
    if st.button("⚙  Procesar archivos", use_container_width=True):
        st.session_state.resultados_por_archivo = {}
        progress = st.progress(0, text="Iniciando…")

        for idx_f, file in enumerate(uploaded_files):
            progress.progress(
                (idx_f) / len(uploaded_files),
                text=f"Procesando {file.name}…"
            )
            try:
                resultados = procesar_pdf(file)
                if resultados:
                    # Generar bytes de Excel para cada cuenta
                    for res in resultados:
                        res["excel_bytes"] = crear_excel(
                            res["info"], res["movimientos"], res["moneda"]
                        )
                    st.session_state.resultados_por_archivo[file.name] = {
                        "ok": True,
                        "resultados": resultados,
                    }
                else:
                    st.session_state.resultados_por_archivo[file.name] = {
                        "ok": False,
                        "msg": "No se detectaron cuentas con movimientos.",
                    }
            except Exception as e:
                st.session_state.resultados_por_archivo[file.name] = {
                    "ok": False,
                    "msg": str(e),
                }

        progress.progress(1.0, text="¡Listo!")

    # ── Resultados ────────────────────────────────────────────────────────────
    for file in uploaded_files:
        nombre = file.name
        if nombre not in st.session_state.resultados_por_archivo:
            continue

        datos = st.session_state.resultados_por_archivo[nombre]

        if not datos["ok"]:
            st.warning(f"**{nombre}** — {datos['msg']}")
            continue

        resultados = datos["resultados"]
        st.markdown(
            f"<div style='font-size:0.72rem;color:#aaa;margin:18px 0 8px;"
            f"font-family:IBM Plex Mono,monospace;letter-spacing:.5px'>"
            f"ARCHIVO · {nombre} · {len(resultados)} cuenta(s) detectada(s)"
            f"</div>",
            unsafe_allow_html=True,
        )

        for res in resultados:
            info   = res["info"]
            movs   = res["movimientos"]
            moneda = res["moneda"]

            # Calcular estadísticas
            movs_reales = [m for m in movs if m["categoria"] != "Saldo Inicial"]
            total_deb   = sum(m["debito"]  or 0 for m in movs_reales)
            total_cred  = sum(m["credito"] or 0 for m in movs_reales)
            saldo_ini   = movs[0]["saldo"] if movs and movs[0]["categoria"] == "Saldo Inicial" else None
            saldo_fin   = movs[-1]["saldo"] if movs else None

            sim         = moneda
            badge_cls   = "result-badge-usd" if moneda == "U$S" else "result-badge"
            nro_display = info.get("nro_cuenta", "—")
            cbu_display = info.get("cbu", "—")

            def fmt_num(v):
                if v is None: return "—"
                return f"{sim} {v:,.2f}"

            st.markdown(f"""
            <div class="result-card">
                <div class="result-card-header">
                    <div>
                        <div class="result-account">Cta Nº {nro_display}</div>
                        <div class="result-meta">CBU: {cbu_display}</div>
                    </div>
                    <span class="{badge_cls}">{moneda}</span>
                </div>
                <div class="stat-row">
                    <div class="stat-item">
                        <div class="stat-label">Movimientos</div>
                        <div class="stat-value">{len(movs_reales)}</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-label">Total débitos</div>
                        <div class="stat-value red">{fmt_num(total_deb)}</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-label">Total créditos</div>
                        <div class="stat-value green">{fmt_num(total_cred)}</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-label">Saldo final</div>
                        <div class="stat-value">{fmt_num(saldo_fin)}</div>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            # Nombre del archivo Excel
            periodo = f"{info.get('desde','').replace('/','')}-{info.get('hasta','').replace('/','')}"
            nro_safe = nro_display.replace("/", "-")
            nombre_xlsx = f"Cta_{nro_safe}_{periodo}.xlsx"

            st.download_button(
                label=f"⬇  Descargar {nombre_xlsx}",
                data=res["excel_bytes"],
                file_name=nombre_xlsx,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_{nombre}_{nro_display}",
                use_container_width=True,
            )

    # ── Leyenda de colores ────────────────────────────────────────────────────
    if any(d["ok"] for d in st.session_state.resultados_por_archivo.values()):
        with st.expander("📋  Leyenda de colores del Excel"):
            st.markdown("""
            <div class="legend-row">
                <span><span class="legend-dot" style="background:#FFF0F0"></span>Débito</span>
                <span><span class="legend-dot" style="background:#F0FFF0"></span>Crédito</span>
                <span><span class="legend-dot" style="background:#FFF8E1"></span>Imp. débitos/créditos · IVA · IIBB</span>
                <span><span class="legend-dot" style="background:#EEF2FF"></span>Saldo inicial</span>
                <span><span class="legend-dot" style="background:#FFF3CD"></span>Sin fecha detectada</span>
            </div>
            """, unsafe_allow_html=True)

